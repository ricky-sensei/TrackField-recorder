import openpyxl


class GetRecord:
    
    def __init__(self):
        self.school_record = []
        

    # pythonでは全角スペースを認識できない。
    def replace_char(self, row, col, ds, filename) -> str:
        if col == 100:
            replaced_char =  "※リレー走者を記入"
        else:
            original_char = r"{}".format(str(ds.cell(row, col).value)).strip()

            if "ボール" in filename or "幅" in filename or "高" in filename:
                replaced_char = original_char.replace(".", "m")
            else:
                replaced_char = original_char.replace(":", "分").replace(".", "秒")

        return replaced_char

    def replace_filename(self, filename: str):
        replaced_filename = filename.split("\\")[1][:-5]
        if "\u3000" in replaced_filename:
            replaced_filename = replaced_filename.replace("\u3000", "　")
        return replaced_filename

    def get_row_number(self, files, selected_school):
        athlete_name = 100
        for filename in files:
            wb = openpyxl.load_workbook(filename)
            data_sheet = wb["Sheet1"]
            for col in range(1, 20):
                cell_value = data_sheet.cell(row=11, column=col).value
                if cell_value == "順位":
                    rank = col
                elif cell_value == "競技者名":
                    athlete_name = col
                elif cell_value in ("所属", "チーム名"):
                    school_name = col
                elif cell_value == "記録":
                    record = col
            for row in data_sheet.iter_rows(min_col=school_name, max_col=school_name):
                for cell in row:
                    rank_value = data_sheet.cell(row=cell.row, column=rank).value
                    if (
                        cell.value == selected_school
                        and isinstance(rank_value, int)
                        and rank_value <= 6
                    ):
                        self.school_record.append(
                            [
                                self.replace_char(cell.row, rank, data_sheet, filename),
                                self.replace_char(cell.row, school_name, data_sheet, filename)+"小学校",
                                self.replace_filename(filename),
                                self.replace_char(cell.row, athlete_name, data_sheet, filename),
                                self.replace_char(cell.row, record, data_sheet, filename),
                            ]
                        )

        return self.school_record
